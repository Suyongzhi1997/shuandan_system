import datetime
import os
import time

from openpyxl.drawing.image import Image
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse, JsonResponse
from django.shortcuts import render, redirect

from django.views.generic.base import View
import openpyxl
from ratelimit.mixins import RatelimitMixin

from delivery.models import Goods, UserFreight, LogisticsCompanyFreight
from files.forms import UploadExcelForm
from files.models import File
from records.models import Record, UserBrushMoney
from records.utils.util import random_str
from shuadan_system.settings import MEDIA_ROOT


class UploadExcelView(LoginRequiredMixin, View):
    def post(self, request):
        # 上传的文件都在request.FILES里面获取，所以这里要多传一个这个参数
        excel_form = UploadExcelForm(request.POST, request.FILES)
        next_url = request.GET.get('next', '')
        if excel_form.is_valid():
            excel_file = excel_form.cleaned_data["excel_file"]
            case_file = File()
            case_file.excel_file = excel_file
            case_file.save()

            excel_file_name = os.path.join(MEDIA_ROOT, case_file.excel_file.name)
            wb = openpyxl.load_workbook(excel_file_name)
            sheet = wb.get_active_sheet()
            for i in range(1, sheet.max_row):
                record_id = sheet.cell(row=i + 1, column=1).value
                record = Record.objects.get(pk=record_id)
                if record:
                    record.site = sheet.cell(row=i + 1, column=2).value
                    record.shop = sheet.cell(row=i + 1, column=3).value
                    record.product_chinese_name = sheet.cell(row=i + 1, column=4).value
                    record.sku = sheet.cell(row=i + 1, column=5).value
                    record.order_word = sheet.cell(row=i + 1, column=6).value
                    record.key_word_page_number = sheet.cell(row=i + 1, column=7).value
                    record.key_word_link = sheet.cell(row=i + 1, column=8).value
                    record.order_date = sheet.cell(row=i + 1, column=9).value
                    record.review_date = sheet.cell(row=i + 1, column=10).value
                    record.direct_review_title = sheet.cell(row=i + 1, column=11).value
                    record.direct_review_content = sheet.cell(row=i + 1, column=12).value
                    record.direct_review_remark = sheet.cell(row=i + 1, column=13).value
                    record.user = request.user
                    record.save()

            messages.add_message(request, messages.SUCCESS, "文件上传成功", extra_tags='success')
            return redirect(next_url)
        else:
            messages.add_message(request, messages.ERROR, "表单验证错误", extra_tags='danger')
            return redirect(next_url)


class UploadCheckExcelView(LoginRequiredMixin, View):
    def post(self, request):
        # 上传的文件都在request.FILES里面获取，所以这里要多传一个这个参数
        excel_form = UploadExcelForm(request.POST, request.FILES)
        next_url = request.GET.get('next', '')
        if excel_form.is_valid():
            excel_file = excel_form.cleaned_data["excel_file"]
            case_file = File()
            case_file.excel_file = excel_file
            case_file.save()

            excel_file_name = os.path.join(MEDIA_ROOT, case_file.excel_file.name)
            wb = openpyxl.load_workbook(excel_file_name)
            sheet = wb.get_active_sheet()
            for i in range(1, sheet.max_row):
                record = Record()

                record.asin = sheet.cell(row=i + 1, column=1).value
                record.c_price = sheet.cell(row=i + 1, column=2).value
                record.purchase_cost = sheet.cell(row=i + 1, column=3).value
                record.product_profit = sheet.cell(row=i + 1, column=4).value
                record.product_upload_time = sheet.cell(row=i + 1, column=5).value
                # record.brush_number = sheet.cell(row=i + 1, column=6).value
                record.sale_30_number = sheet.cell(row=i + 1, column=6).value
                record.sale_7_number = sheet.cell(row=i + 1, column=7).value
                record.record_src = sheet.cell(row=i + 1, column=8).value
                record.now_score = sheet.cell(row=i + 1, column=9).value
                record.now_review_number = sheet.cell(row=i + 1, column=10).value
                review_type = sheet.cell(row=i + 1, column=11).value
                record.review_type = review_type
                if review_type == "留评":
                    record.direct_review = '0'
                    record.free_review_number = '0'
                    record.feedback = '0'
                elif review_type == "直评":
                    record.review_number = '0'
                    record.free_review_number = '0'
                    record.feedback = '0'
                elif review_type == "免评":
                    record.review_number = '0'
                    record.direct_review = '0'
                    record.feedback = '0'
                elif review_type == "feedback":
                    record.review_number = '0'
                    record.direct_review = '0'
                    record.free_review_number = '0'

                record.user = request.user
                record.save()

            messages.add_message(request, messages.SUCCESS, "文件上传成功", extra_tags='success')
            return redirect(next_url)
        else:
            messages.add_message(request, messages.ERROR, "表单验证错误", extra_tags='danger')
            return redirect(next_url)


class UploadDeliveryExcelView(LoginRequiredMixin, View):
    def post(self, request):
        # 上传的文件都在request.FILES里面获取，所以这里要多传一个这个参数
        excel_form = UploadExcelForm(request.POST, request.FILES)
        next_url = request.GET.get('next', '')
        if excel_form.is_valid():
            excel_file = excel_form.cleaned_data["excel_file"]
            case_file = File()
            case_file.excel_file = excel_file
            case_file.save()

            excel_file_name = os.path.join(MEDIA_ROOT, case_file.excel_file.name)
            wb = openpyxl.load_workbook(excel_file_name)
            sheet = wb.get_active_sheet()
            for i in range(1, sheet.max_row):
                track_number = sheet.cell(row=i + 1, column=6).value
                if not Goods.objects.filter(user=request.user, track_number=track_number):
                    good = Goods()
                    old_total_price = good.total_price if good.total_price else 0
                    good.track_number = track_number
                    good.user = request.user
                    good.send_time = sheet.cell(row=i + 1, column=1).value
                    if isinstance(good.send_time, str):
                        good.send_time = datetime.datetime.strptime(good.send_time.replace("/", "-"), "%Y-%m-%d")
                    good.send_company = sheet.cell(row=i + 1, column=2).value
                    good.channel = sheet.cell(row=i + 1, column=3).value
                    good.site = sheet.cell(row=i + 1, column=4).value
                    good.country = sheet.cell(row=i + 1, column=5).value
                    good.net_weight = sheet.cell(row=i + 1, column=7).value
                    good.volume_weight = sheet.cell(row=i + 1, column=8).value
                    good.actual_charged_weight = sheet.cell(row=i + 1, column=9).value
                    good.pieces_number = sheet.cell(row=i + 1, column=10).value
                    good.includes_price = sheet.cell(row=i + 1, column=11).value
                    good.other_price = sheet.cell(row=i + 1, column=12).value
                    good.total_price = sheet.cell(row=i + 1, column=13).value
                    good.express_time = sheet.cell(row=i + 1, column=14).value
                    good.express_status = sheet.cell(row=i + 1, column=15).value if sheet.cell(row=i + 1,
                                                                                               column=15).value else "not express"
                    good.settlement = sheet.cell(row=i + 1, column=16).value
                    good.order_remark = sheet.cell(row=i + 1, column=17).value
                    good.user = request.user
                    good.save()

                    total_price = sheet.cell(row=i + 1, column=13).value
                    send_company = sheet.cell(row=i + 1, column=2).value
                    good_total_price = good.total_price if good.total_price else 0

                    year = datetime.datetime.now().strftime('%Y')
                    month = datetime.datetime.now().strftime('%m')
                    user_freight_money = UserFreight.objects.filter(brush_year=year, brush_month=month,
                                                                    user=request.user).first()
                    total_price_difference = float(good_total_price) - float(old_total_price)
                    if not user_freight_money:
                        user_freight_money_object = UserFreight()
                        user_freight_money_object.user = request.user
                    else:
                        user_freight_money_object = user_freight_money

                    user_freight_money_object.freight += total_price_difference
                    user_freight_money_object.save()

                    company_freight_money = LogisticsCompanyFreight.objects.filter(brush_year=year, brush_month=month,
                                                                                   send_company=send_company).first()
                    if not company_freight_money:
                        company_freight_money_object = LogisticsCompanyFreight()
                        company_freight_money_object.send_company = send_company
                    else:
                        company_freight_money_object = company_freight_money
                    company_freight_money_object.freight += total_price_difference
                    company_freight_money_object.save()

            messages.add_message(request, messages.SUCCESS, "文件上传成功", extra_tags='success')
            return redirect(next_url)
        else:
            messages.add_message(request, messages.ERROR, "表单验证错误", extra_tags='danger')
            return redirect(next_url)


class DownloadExcelView(LoginRequiredMixin, View):
    def get(self, request):
        id_data = request.GET.get('id_data')
        record_id_list = [i for i in id_data.split('_') if i != '']
        wb = openpyxl.Workbook()
        sheet = wb.get_active_sheet()
        sheet.freeze_panes = 'A2'
        sheet.append(
            ["ASIN", "运营人员", "产品售价", "采购成本", "产品利润", "上架时间", "产品刷单总数量", "最近30天销量", "最近7天销量", "留评", "Feedback", "直评",
             "免评",
             "站点", "店铺", "产品中文名", "SKU", "刷单词", "关键词的页数", "主图片", "关键词链接", "刷单公司", "订单号", "刷单费用", "留评反馈", "提交日期",
             "订单状态", "直评标题", "直评内容", "直评备注", "下单日期", "上评日期"])

        i = 2
        for record_id in record_id_list:
            record_object = Record.objects.get(pk=record_id)
            sheet['A{}'.format(i)] = record_object.asin
            sheet['B{}'.format(
                i)] = record_object.user.nick_name if record_object.user.nick_name else record_object.user.username
            sheet['C{}'.format(i)] = record_object.c_price
            sheet['D{}'.format(i)] = record_object.purchase_cost
            sheet['E{}'.format(i)] = record_object.product_profit
            sheet['F{}'.format(i)] = record_object.product_upload_time.strftime('%Y-%m-%d')
            sheet['G{}'.format(i)] = record_object.brush_number
            sheet['H{}'.format(i)] = record_object.sale_30_number
            sheet['I{}'.format(i)] = record_object.sale_7_number
            sheet['J{}'.format(i)] = record_object.review_number
            sheet['K{}'.format(i)] = record_object.feedback
            sheet['L{}'.format(i)] = record_object.direct_review
            sheet['M{}'.format(i)] = record_object.free_review_number
            sheet['N{}'.format(i)] = record_object.site
            sheet['O{}'.format(i)] = record_object.shop
            sheet['P{}'.format(i)] = record_object.product_chinese_name
            sheet['Q{}'.format(i)] = record_object.sku
            sheet['R{}'.format(i)] = record_object.order_word
            sheet['S{}'.format(i)] = record_object.key_word_page_number
            # 插入图片
            if record_object.main_image:
                img_file = os.path.join(MEDIA_ROOT, record_object.main_image.name)
                img = Image(img_file)
                sheet.column_dimensions['T'].width = 15.0
                sheet.row_dimensions[i].height = 80
                img.width = 100.0
                img.height = 100.0
                sheet.add_image(img, 'T{}'.format(i))
            else:
                sheet['T{}'.format(i)] = '暂无图片'

            sheet['U{}'.format(i)] = record_object.key_word_link
            if record_object.brush_company:
                sheet['V{}'.format(
                    i)] = record_object.brush_company.nick_name if record_object.brush_company.nick_name else record_object.brush_company.username
            else:
                sheet['V{}'.format(i)] = ""

            sheet['W{}'.format(i)] = record_object.order_number
            sheet['X{}'.format(i)] = record_object.brush_money
            sheet['Y{}'.format(i)] = record_object.review_feedback
            sheet['Z{}'.format(i)] = record_object.add_time.strftime('%Y-%m-%d')
            sheet['AA{}'.format(i)] = record_object.record_status
            sheet['AB{}'.format(i)] = record_object.direct_review_title
            sheet['AC{}'.format(i)] = record_object.direct_review_content
            sheet['AD{}'.format(i)] = record_object.direct_review_remark
            sheet['AE{}'.format(i)] = record_object.order_date.strftime('%Y-%m-%d') if record_object.order_date else ""
            sheet['AF{}'.format(i)] = record_object.review_date.strftime(
                '%Y-%m-%d') if record_object.review_date else ""

            i += 1

        today = time.strftime("%Y-%m-%d", time.localtime())
        file_path = os.path.join(MEDIA_ROOT, 'download', today)
        if not os.path.exists(file_path):
            os.makedirs(file_path)

        random_name = random_str(16)
        file_name = os.path.join(MEDIA_ROOT, 'download', today, random_name + '.xlsx')
        wb.save(file_name)

        data = {
            "file_name": os.path.join('media', 'download', today, random_name + '.xlsx')
        }
        return JsonResponse(data)


class DownloadBrushMoneyExcelView(LoginRequiredMixin, View):
    def get(self, request):
        id_data = request.GET.get('id_data')
        id_list = [i for i in id_data.split('_') if i != '']
        wb = openpyxl.Workbook()
        sheet = wb.get_active_sheet()
        sheet.freeze_panes = 'A2'
        sheet.append(["年", "月", "费用", "运营人员"])

        for i in id_list:
            brush_money = UserBrushMoney.objects.get(pk=i)
            username = brush_money.user.nick_name if brush_money.user.nick_name else brush_money.user.username
            sheet.append([brush_money.brush_year, brush_money.brush_month, brush_money.brush_money, username])

        today = time.strftime("%Y-%m-%d", time.localtime())
        file_path = os.path.join(MEDIA_ROOT, 'download', today)
        if not os.path.exists(file_path):
            os.makedirs(file_path)

        random_name = random_str(16)
        file_name = os.path.join(MEDIA_ROOT, 'download', today, random_name + '.xlsx')
        wb.save(file_name)

        data = {
            "file_name": os.path.join('media', 'download', today, random_name + '.xlsx')
        }
        return JsonResponse(data)


class DownloadInfoView(LoginRequiredMixin, View):
    def get(self, request):
        id_data = request.GET.get('id_data')
        id_list = [i for i in id_data.split('_') if i != '']
        wb = openpyxl.Workbook()
        sheet = wb.get_active_sheet()
        sheet.freeze_panes = 'A2'
        sheet.append(
            ["发货日期", "发货公司", "渠道", "站点", "运营人员", "国家", "运营追踪号", "实重", "材积重", "实际收费重量", "件数", "单价包含纺织等的附加费", "其他附加费用",
             "总运费", "到货日期", "是否全部到货", "结算", "其他备注"])

        for i in id_list:
            info_object = Goods.objects.get(pk=i)
            username = info_object.user.nick_name if info_object.user.nick_name else info_object.user.username
            send_time = info_object.send_time.strftime('%Y-%m-%d') if info_object.send_time else ""
            express_time = info_object.express_time.strftime('%Y-%m-%d') if info_object.express_time else ""
            sheet.append([send_time, info_object.send_company, info_object.channel,
                          info_object.site,
                          username, info_object.country, info_object.track_number, info_object.net_weight,
                          info_object.volume_weight, info_object.actual_charged_weight, info_object.pieces_number,
                          info_object.includes_price, info_object.other_price, info_object.total_price,
                          express_time, info_object.get_express_status_display(),
                          info_object.settlement, info_object.order_remark])

        today = time.strftime("%Y-%m-%d", time.localtime())
        file_path = os.path.join(MEDIA_ROOT, 'download', today)
        if not os.path.exists(file_path):
            os.makedirs(file_path)

        random_name = random_str(16)
        file_name = os.path.join(MEDIA_ROOT, 'download', today, random_name + '.xlsx')
        wb.save(file_name)

        data = {
            "file_name": os.path.join('media', 'download', today, random_name + '.xlsx')
        }
        return JsonResponse(data)


class DownloadAllInfoView(LoginRequiredMixin, View):
    def get(self, request):
        # id_data = request.GET.get('id_data')
        # id_list = [i for i in id_data.split('_') if i != '']
        wb = openpyxl.Workbook()
        sheet = wb.get_active_sheet()
        sheet.freeze_panes = 'A2'
        sheet.append(
            ["发货日期", "发货公司", "渠道", "站点", "运营人员", "国家", "运营追踪号", "实重", "材积重", "实际收费重量", "件数", "单价包含纺织等的附加费", "其他附加费用",
             "总运费", "到货日期", "是否全部到货", "结算", "其他备注"])
        infos = Goods.objects.all()
        for info_object in infos:
            username = info_object.user.nick_name if info_object.user.nick_name else info_object.user.username
            send_time = info_object.send_time.strftime('%Y-%m-%d') if info_object.send_time else ""
            express_time = info_object.express_time.strftime('%Y-%m-%d') if info_object.express_time else ""
            sheet.append([send_time, info_object.send_company, info_object.channel,
                          info_object.site,
                          username, info_object.country, info_object.track_number, info_object.net_weight,
                          info_object.volume_weight, info_object.actual_charged_weight, info_object.pieces_number,
                          info_object.includes_price, info_object.other_price, info_object.total_price,
                          express_time, info_object.get_express_status_display(),
                          info_object.settlement, info_object.order_remark])

        today = time.strftime("%Y-%m-%d", time.localtime())
        file_path = os.path.join(MEDIA_ROOT, 'download', today)
        if not os.path.exists(file_path):
            os.makedirs(file_path)

        random_name = random_str(16)
        file_name = os.path.join(MEDIA_ROOT, 'download', today, random_name + '.xlsx')
        wb.save(file_name)

        data = {
            "file_name": os.path.join('media', 'download', today, random_name + '.xlsx')
        }
        return JsonResponse(data)


class DownloadAllRecordExcelView(LoginRequiredMixin, View):
    def get(self, request):
        wb = openpyxl.Workbook()
        sheet = wb.get_active_sheet()
        sheet.freeze_panes = 'A2'
        sheet.append(
            ["ASIN", "运营人员", "产品售价", "采购成本", "产品利润", "上架时间", "产品刷单总数量", "最近30天销量", "最近7天销量", "留评", "Feedback", "直评",
             "免评",
             "站点", "店铺", "产品中文名", "SKU", "刷单词", "关键词的页数", "主图片", "关键词链接", "刷单公司", "订单号", "刷单费用", "留评反馈", "提交日期",
             "订单状态", "直评标题", "直评内容", "直评备注", "下单日期", "上评日期"])

        i = 2
        record_objects = Record.objects.all()
        for record_object in record_objects:

            sheet['A{}'.format(i)] = record_object.asin
            sheet['B{}'.format(
                i)] = record_object.user.nick_name if record_object.user.nick_name else record_object.user.username
            sheet['C{}'.format(i)] = record_object.c_price
            sheet['D{}'.format(i)] = record_object.purchase_cost
            sheet['E{}'.format(i)] = record_object.product_profit
            sheet['F{}'.format(i)] = record_object.product_upload_time.strftime('%Y-%m-%d')
            sheet['G{}'.format(i)] = record_object.brush_number
            sheet['H{}'.format(i)] = record_object.sale_30_number
            sheet['I{}'.format(i)] = record_object.sale_7_number
            sheet['J{}'.format(i)] = record_object.review_number
            sheet['K{}'.format(i)] = record_object.feedback
            sheet['L{}'.format(i)] = record_object.direct_review
            sheet['M{}'.format(i)] = record_object.free_review_number
            sheet['N{}'.format(i)] = record_object.site
            sheet['O{}'.format(i)] = record_object.shop
            sheet['P{}'.format(i)] = record_object.product_chinese_name
            sheet['Q{}'.format(i)] = record_object.sku
            sheet['R{}'.format(i)] = record_object.order_word
            sheet['S{}'.format(i)] = record_object.key_word_page_number
            # 插入图片
            if record_object.main_image:
                img_file = os.path.join(MEDIA_ROOT, record_object.main_image.name)
                img = Image(img_file)
                sheet.column_dimensions['T'].width = 15.0
                sheet.row_dimensions[i].height = 80
                img.width = 100.0
                img.height = 100.0
                sheet.add_image(img, 'T{}'.format(i))
            else:
                sheet['T{}'.format(i)] = '暂无图片'

            sheet['U{}'.format(i)] = record_object.key_word_link
            if record_object.brush_company:
                sheet['V{}'.format(
                    i)] = record_object.brush_company.nick_name if record_object.brush_company.nick_name else record_object.brush_company.username
            else:
                sheet['V{}'.format(i)] = ""

            sheet['W{}'.format(i)] = record_object.order_number
            sheet['X{}'.format(i)] = record_object.brush_money
            sheet['Y{}'.format(i)] = record_object.review_feedback
            sheet['Z{}'.format(i)] = record_object.add_time.strftime('%Y-%m-%d')
            sheet['AA{}'.format(i)] = record_object.record_status
            sheet['AB{}'.format(i)] = record_object.direct_review_title
            sheet['AC{}'.format(i)] = record_object.direct_review_content
            sheet['AD{}'.format(i)] = record_object.direct_review_remark
            sheet['AE{}'.format(i)] = record_object.order_date.strftime('%Y-%m-%d') if record_object.order_date else ""
            sheet['AF{}'.format(i)] = record_object.review_date.strftime(
                '%Y-%m-%d') if record_object.review_date else ""

            i += 1

        today = time.strftime("%Y-%m-%d", time.localtime())
        file_path = os.path.join(MEDIA_ROOT, 'download', today)
        if not os.path.exists(file_path):
            os.makedirs(file_path)

        random_name = random_str(16)
        file_name = os.path.join(MEDIA_ROOT, 'download', today, random_name + '.xlsx')
        wb.save(file_name)

        data = {
            "file_name": os.path.join('media', 'download', today, random_name + '.xlsx')
        }
        return JsonResponse(data)


class DownloadTemplateView(LoginRequiredMixin, View):
    def get(self, request):
        filename = request.GET.get('filename', '')
        data = {
            "file_name": os.path.join('media', 'download', filename)
        }
        return JsonResponse(data)


class UploadDirectReviewFileView(LoginRequiredMixin, View):
    def post(self, request):
        # 上传的文件都在request.FILES里面获取，所以这里要多传一个这个参数
        excel_form = UploadExcelForm(request.POST, request.FILES)
        next_url = request.GET.get('next', '')
        record_id = request.POST.get('record_id')
        if excel_form.is_valid():
            record = Record.objects.get(pk=record_id)
            excel_file = excel_form.cleaned_data["excel_file"]
            excel_file.name = random_str(16) + '.xlsx'
            record.direct_review_file = excel_file

            record.save()
            messages.add_message(request, messages.SUCCESS, "文件上传成功", extra_tags='success')
            return redirect(next_url)
        else:
            messages.add_message(request, messages.ERROR, "表单验证错误", extra_tags='danger')
            return redirect(next_url)


class SdgsDownloadView(RatelimitMixin, LoginRequiredMixin, View):
    ratelimit_key = 'ip'
    ratelimit_rate = '5/d'
    ratelimit_block = True
    ratelimit_method = 'GET'

    def get(self, request):
        id_data = request.GET.get('id_data')
        record_id_list = [i for i in id_data.split('_') if i != '']
        wb = openpyxl.Workbook()
        sheet = wb.get_active_sheet()
        sheet.freeze_panes = 'A2'
        sheet.append(
            ["ASIN", "产品售价", "留评", "Feedback", "直评", "免评", "国家", "店铺", "产品中文名", "SKU", "刷单词", "关键词的页数", "主图片", "关键词链接",
             "直评标题", "直评内容", "直评备注", "下单日期", "上评日期"])

        i = 2
        for record_id in record_id_list:
            record_object = Record.objects.get(pk=record_id)
            sheet['A{}'.format(i)] = record_object.asin
            sheet['B{}'.format(i)] = record_object.c_price
            sheet['C{}'.format(i)] = record_object.review_number
            sheet['D{}'.format(i)] = record_object.feedback
            sheet['E{}'.format(i)] = record_object.direct_review
            sheet['F{}'.format(i)] = record_object.free_review_number
            sheet['G{}'.format(i)] = record_object.site
            sheet['H{}'.format(i)] = record_object.shop
            sheet['I{}'.format(i)] = record_object.product_chinese_name
            sheet['J{}'.format(i)] = record_object.sku
            sheet['K{}'.format(i)] = record_object.order_word
            sheet['L{}'.format(i)] = record_object.key_word_page_number
            # 插入图片
            if record_object.main_image:
                img_file = os.path.join(MEDIA_ROOT, record_object.main_image.name)
                img = Image(img_file)
                sheet.column_dimensions['M'].width = 15.0
                sheet.row_dimensions[i].height = 80
                img.width = 100.0
                img.height = 100.0
                sheet.add_image(img, 'M{}'.format(i))
            else:
                sheet['M{}'.format(i)] = '暂无图片'

            sheet['N{}'.format(i)] = record_object.key_word_link
            sheet['O{}'.format(i)] = record_object.direct_review_title
            sheet['P{}'.format(i)] = record_object.direct_review_content
            sheet['Q{}'.format(i)] = record_object.direct_review_remark
            sheet['R{}'.format(i)] = record_object.order_date.strftime('%Y-%m-%d') if record_object.order_date else ""
            sheet['S{}'.format(i)] = record_object.review_date.strftime('%Y-%m-%d') if record_object.review_date else ""

            i += 1

        today = time.strftime("%Y-%m-%d", time.localtime())
        file_path = os.path.join(MEDIA_ROOT, 'download', today)
        if not os.path.exists(file_path):
            os.makedirs(file_path)

        random_name = random_str(16)
        file_name = os.path.join(MEDIA_ROOT, 'download', today, random_name + '.xlsx')
        wb.save(file_name)

        data = {
            "file_name": os.path.join('media', 'download', today, random_name + '.xlsx')
        }
        return JsonResponse(data)
